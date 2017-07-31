#!/usr/bin/python
# -*- coding: UTF-8 -*-

import sys
import datetime
from openpyxl import load_workbook

wb = load_workbook('voice.xlsx')

class voiceCalculator():
	def __init__(self):
		total_mins = 0;
		ws = wb["语音清单"]
		for row in ws.iter_rows():
			bNeedCalculate = 0
			for cell in row:
				if 'C' in cell.coordinate:
					if cell.value and '主叫' in cell.value:
						bNeedCalculate = 1
				elif 'E' in cell.coordinate:
					if bNeedCalculate == 1:
						dt = datetime.datetime.strptime('2017-07-31 %s' % (cell.value), '%Y-%m-%d %H:%M:%S')
						cur_mins = dt.hour * 60 + dt.minute + 1
						total_mins = total_mins + cur_mins
		print("Total minutes : %s" % total_mins)				
				
if(__name__ == "__main__"):
	voiceCalculator()
	